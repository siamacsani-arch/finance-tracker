#!/usr/bin/env python3
"""
db.py — SQLite helper for the Finance Dashboard
"""
import sqlite3
from pathlib import Path
from datetime import datetime

BASE = Path(__file__).parent
DB_PATH = BASE / "transactions.db"

BUDGETS_DEFAULT = {
    "Work": 3100, "Clients": 200, "Tax Return": 0, "Friends": 0,
    "Rent": 3075, "Groceries": 300, "Subway": 80, "Transfers": 200,
    "Fast Food": 80, "DoorDash": 100, "Furniture": 50, "Clothes": 100,
    "Bars": 100, "Movies": 50, "Dates": 300, "Video Games": 20,
    "Personal": 100, "Eventus": 100, "Streaming": 20,
    "Student Loans": 324, "Uber": 50, "Coffee": 30, "Celcius": 20,
    "Internet": 50, "Entertainment": 50, "Laundry": 20,
    "Amazon": 100, "Social": 30, "AI": 50, "Haircut": 50,
    "Eyebrows": 15, "Handyman": 50, "Toys": 50, "Misc.": 30,
    "Clubs": 50, "Shows": 50, "Tax Guy": 0, "Gas": 50, "Tax Returns": 0,
    "Health": 50, "Temu": 50, "Business": 150,
}

INCOME_CATEGORIES = {"Work", "Clients", "Tax Return", "Friends", "Tax Returns"}


def get_conn():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT NOT NULL,
            type        TEXT NOT NULL,
            category    TEXT NOT NULL,
            amount      REAL NOT NULL,
            description TEXT,
            source      TEXT DEFAULT 'excel'
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS budgets (
            category        TEXT PRIMARY KEY,
            type            TEXT NOT NULL,
            monthly_budget  REAL NOT NULL DEFAULT 0
        )
    """)
    conn.commit()
    conn.close()


def populate_budgets():
    conn = get_conn()
    c = conn.cursor()
    for category, amount in BUDGETS_DEFAULT.items():
        btype = "Income" if category in INCOME_CATEGORIES else "Expense"
        c.execute("""
            INSERT OR IGNORE INTO budgets (category, type, monthly_budget)
            VALUES (?, ?, ?)
        """, (category, btype, amount))
    conn.commit()
    conn.close()


def get_budgets():
    conn = get_conn()
    rows = conn.execute("SELECT category, type, monthly_budget FROM budgets ORDER BY type DESC, category").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def upsert_budget(category: str, amount: float):
    conn = get_conn()
    btype = "Income" if category in INCOME_CATEGORIES else "Expense"
    conn.execute("""
        INSERT INTO budgets (category, type, monthly_budget)
        VALUES (?, ?, ?)
        ON CONFLICT(category) DO UPDATE SET monthly_budget = excluded.monthly_budget
    """, (category, btype, amount))
    conn.commit()
    conn.close()


def get_summary(year: int, month: int) -> dict:
    conn = get_conn()
    income = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='Income' AND strftime('%Y',date)=? AND strftime('%m',date)=?",
        (str(year), f"{month:02d}")
    ).fetchone()[0]
    expenses = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='Expense' AND strftime('%Y',date)=? AND strftime('%m',date)=?",
        (str(year), f"{month:02d}")
    ).fetchone()[0]
    conn.close()
    balance = income - expenses
    savings_rate = round((balance / income * 100), 1) if income > 0 else 0.0
    return {
        "income": round(income, 2),
        "expenses": round(expenses, 2),
        "balance": round(balance, 2),
        "savings_rate": savings_rate,
    }


def get_transactions(year: int, month: int, txn_type: str = None, category: str = None) -> list:
    conn = get_conn()
    query = "SELECT * FROM transactions WHERE strftime('%Y',date)=? AND strftime('%m',date)=?"
    params = [str(year), f"{month:02d}"]
    if txn_type:
        query += " AND type=?"
        params.append(txn_type)
    if category:
        query += " AND category=?"
        params.append(category)
    query += " ORDER BY date DESC, id DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_trends(year: int) -> list:
    conn = get_conn()
    result = []
    for m in range(1, 13):
        mm = f"{m:02d}"
        income = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='Income' AND strftime('%Y',date)=? AND strftime('%m',date)=?",
            (str(year), mm)
        ).fetchone()[0]
        expenses = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='Expense' AND strftime('%Y',date)=? AND strftime('%m',date)=?",
            (str(year), mm)
        ).fetchone()[0]
        result.append({
            "month": m,
            "income": round(income, 2),
            "expenses": round(expenses, 2),
            "balance": round(income - expenses, 2),
        })
    conn.close()
    return result


def get_spending_by_category(year: int, month: int) -> list:
    conn = get_conn()
    rows = conn.execute(
        """SELECT category, SUM(amount) as amount
           FROM transactions
           WHERE type='Expense'
             AND strftime('%Y',date)=?
             AND strftime('%m',date)=?
           GROUP BY category
           ORDER BY amount DESC""",
        (str(year), f"{month:02d}")
    ).fetchall()
    conn.close()
    return [{"category": r["category"], "amount": round(r["amount"], 2)} for r in rows]


def get_budget_vs_actual(year: int, month: int) -> list:
    conn = get_conn()
    budgets = {r["category"]: r for r in conn.execute("SELECT * FROM budgets").fetchall()}

    # Aggregate actuals
    actuals_expense = {}
    for row in conn.execute(
        "SELECT category, SUM(amount) as total FROM transactions WHERE type='Expense' AND strftime('%Y',date)=? AND strftime('%m',date)=? GROUP BY category",
        (str(year), f"{month:02d}")
    ).fetchall():
        actuals_expense[row["category"]] = round(row["total"], 2)

    actuals_income = {}
    for row in conn.execute(
        "SELECT category, SUM(amount) as total FROM transactions WHERE type='Income' AND strftime('%Y',date)=? AND strftime('%m',date)=? GROUP BY category",
        (str(year), f"{month:02d}")
    ).fetchall():
        actuals_income[row["category"]] = round(row["total"], 2)

    conn.close()

    result = []
    for cat, b in budgets.items():
        if b["type"] == "Income":
            actual = actuals_income.get(cat, 0.0)
        else:
            actual = actuals_expense.get(cat, 0.0)
        budget = b["monthly_budget"]
        remaining = budget - actual
        pct_used = round((actual / budget * 100), 1) if budget > 0 else (100.0 if actual > 0 else 0.0)
        result.append({
            "category": cat,
            "type": b["type"],
            "budget": budget,
            "actual": actual,
            "remaining": round(remaining, 2),
            "pct_used": pct_used,
        })
    return result


def update_transaction_category(txn_id: int, category: str):
    conn = get_conn()
    conn.execute("UPDATE transactions SET category=? WHERE id=?", (category, txn_id))
    conn.commit()
    conn.close()


def insert_transaction(t: dict):
    """Insert a single transaction dict (date, type, category, amount, description, source)."""
    conn = get_conn()
    conn.execute(
        "INSERT INTO transactions (date, type, category, amount, description, source) VALUES (?,?,?,?,?,?)",
        (
            str(t.get("date", "")),
            t.get("type", "Expense"),
            t.get("category", "Misc."),
            float(t.get("amount", 0)),
            t.get("description", t.get("merchant", "")),
            t.get("source", "sync"),
        )
    )
    conn.commit()
    conn.close()


def import_from_excel(xlsx_path: str):
    """Read all transactions from Excel Transactions sheet and upsert into SQLite."""
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path))
    ws = wb["Transactions"]

    conn = get_conn()
    # Clear existing excel-sourced rows and reimport fresh
    conn.execute("DELETE FROM transactions WHERE source='excel'")
    conn.commit()

    inserted = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Columns: row#, date, type, category, amount, description
        if row[1] is None or row[4] is None:
            continue
        date_val = row[1]
        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]

        txn_type = row[2] or "Expense"
        category = row[3] or "Misc."
        amount = float(row[4])
        description = row[5] or ""

        conn.execute(
            "INSERT INTO transactions (date, type, category, amount, description, source) VALUES (?,?,?,?,?,?)",
            (date_str, txn_type, category, amount, description, "excel")
        )
        inserted += 1

    conn.commit()
    conn.close()
    print(f"Imported {inserted} transactions from Excel.")
    return inserted


if __name__ == "__main__":
    print("Initializing database...")
    init_db()
    populate_budgets()
    print("Importing from Excel...")
    xlsx = BASE / "Financial Tracker.xlsx"
    if xlsx.exists():
        import_from_excel(str(xlsx))
    else:
        print(f"Warning: {xlsx} not found, skipping import.")
    print("Done. Database ready at", DB_PATH)
