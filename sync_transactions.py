#!/usr/bin/env python3
"""
Financial Tracker — Gmail Auto-Sync
Reads Wells Fargo + Amex transaction alert emails and appends to Financial Tracker.xlsx
"""
import imaplib
import email
import re
import json
import os
import sys
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook
from html import unescape

# SQLite DB helper (optional — only used if db.py exists alongside this script)
try:
    import db as _db
    _DB_AVAILABLE = True
except ImportError:
    _DB_AVAILABLE = False

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE       = Path(__file__).parent
XLSX_PATH  = BASE / "Financial Tracker.xlsx"
CONFIG     = BASE / "config.json"
PROCESSED  = BASE / ".processed_ids.json"
LOG_PATH   = BASE / "sync.log"

# ── Category keyword map (merchant → category) ────────────────────────────────
CATEGORY_MAP = [
    # Food delivery
    ("DOORDASH",        "DoorDash"),
    ("DOOR DASH",       "DoorDash"),
    ("UBER EATS",       "DoorDash"),
    ("UBEREATS",        "DoorDash"),
    ("GRUBHUB",         "DoorDash"),
    ("SEAMLESS",        "DoorDash"),
    # Groceries
    ("WHOLE FOODS",     "Groceries"),
    ("WHOLEFDS",        "Groceries"),
    ("TRADER JOE",      "Groceries"),
    ("SHOPRITE",        "Groceries"),
    ("FOOD BAZAAR",     "Groceries"),
    ("KEY FOOD",        "Groceries"),
    ("ALDI",            "Groceries"),
    ("WALMART GROCER",  "Groceries"),
    ("STOP & SHOP",     "Groceries"),
    ("FAIRWAY",         "Groceries"),
    # Fast food
    ("UP INC",          "Fast Food"),
    ("77TH STREET",     "Fast Food"),
    ("PRET A MANGER",   "Fast Food"),
    ("PRET",            "Fast Food"),
    ("MCDONALD",        "Fast Food"),
    ("BURGER KING",     "Fast Food"),
    ("WENDY",           "Fast Food"),
    ("TACO BELL",       "Fast Food"),
    ("CHICK-FIL-A",     "Fast Food"),
    ("CHIPOTLE",        "Fast Food"),
    ("SHAKE SHACK",     "Fast Food"),
    ("FIVE GUYS",       "Fast Food"),
    ("POPEYES",         "Fast Food"),
    ("WINGSTOP",        "Fast Food"),
    ("PIZZA",           "Fast Food"),
    ("BAGEL",           "Fast Food"),
    ("DUNKIN",          "Fast Food"),
    ("POKE",            "Fast Food"),
    ("SWEETGREEN",      "Fast Food"),
    # Coffee
    ("STARBUCKS",       "Coffee"),
    ("BLUESTONE",       "Coffee"),
    ("GREGORYS",        "Coffee"),
    # Ride share
    ("UBER",            "Uber"),   # must come AFTER UBER EATS
    ("LYFT",            "Uber"),
    # Subway transit
    ("MTA",             "Subway"),
    ("NYCT",            "Subway"),
    # Streaming / social / Apple subscriptions
    ("INSTAGRAM",       "Streaming"),
    ("IG BADGE",        "Streaming"),
    ("APPLE.COM/BILL",  "Streaming"),
    ("NETFLIX",         "Streaming"),
    # Business / dev tools
    ("APPLE DEVELOPER", "Business"),
    ("APPLECOM",        "Business"),
    ("GITHUB",          "Business"),
    ("NOTION",          "Business"),
    ("ZOOM",            "Business"),
    ("GSUITE",          "Business"),
    ("GOOGLE WORKSPACE","Business"),
    ("SHOPIFY",         "Business"),
    ("NAMECHEAP",       "Business"),
    ("GODADDY",         "Business"),
    ("DIGITALOCEAN",    "Business"),
    ("AWS",             "Business"),
    ("HEROKU",          "Business"),
    ("SPOTIFY",         "Streaming"),
    ("HULU",            "Streaming"),
    ("DISNEY+",         "Streaming"),
    ("HBO",             "Streaming"),
    ("APPLE.COM/BILL",  "Streaming"),
    ("YOUTUBE PREMIUM", "Streaming"),
    # AI
    ("OPEN AI",         "AI"),
    ("OPENAI",          "AI"),
    ("CHATGPT",         "AI"),
    ("MIDJOURNEY",      "AI"),
    ("CURSOR",          "AI"),
    ("REPLICATE",       "AI"),
    ("ELEVEN LABS",     "AI"),
    ("ELEVENLABS",      "AI"),
    ("LUMA AI",         "AI"),
    ("LUMALABS",        "AI"),
    ("ANTHROPIC",       "AI"),
    # Amazon
    ("AMAZON",          "Amazon"),
    ("AMZN",            "Amazon"),
    # Temu
    ("TEMU",            "Temu"),
    # Health
    ("NEXUS RX",        "Health"),
    ("NEXUSRX",         "Health"),
    ("ADVANCED DERM",   "Health"),
    ("DERMATOLOGY",     "Health"),
    ("CVS",             "Health"),
    ("WALGREENS",       "Health"),
    ("PHARMACY",        "Health"),
    # Clothes
    ("ZARA",            "Clothes"),
    ("H&M",             "Clothes"),
    ("UNIQLO",          "Clothes"),
    ("GAP",             "Clothes"),
    ("NIKE",            "Clothes"),
    ("ADIDAS",          "Clothes"),
    ("ASOS",            "Clothes"),
    # Entertainment
    ("AMC THEATRES",    "Movies"),
    ("AMC THEATRE",     "Movies"),
    ("REGAL",           "Movies"),
    ("CINEMARK",        "Movies"),
    ("ALAMO",           "Movies"),
    ("TICKETMASTER",    "Shows"),
    ("STUBHUB",         "Shows"),
    ("EVENTBRITE",      "Shows"),
    # Dating apps
    ("HINGE",           "Social"),
    ("TINDER",          "Social"),
    ("BUMBLE",          "Social"),
    ("MATCH.COM",       "Social"),
    # Student loans
    ("NAVIENT",         "Student Loans"),
    ("MOHELA",          "Student Loans"),
    ("GREAT LAKES",     "Student Loans"),
    ("NELNET",          "Student Loans"),
    ("EDFINANCIAL",     "Student Loans"),
    # Internet / bills
    ("SPECTRUM",        "Internet"),
    ("VERIZON",         "Internet"),
    ("AT&T",            "Internet"),
    ("XFINITY",         "Internet"),
    # Personal
    ("VAPE",            "Personal"),
    ("JUUL",            "Personal"),
    ("TOBACCO",         "Personal"),
    ("DISPENSARY",      "Personal"),
    ("CANNABIS",        "Personal"),
    ("STIIIZY",         "Personal"),
    # Bars / clubs
    ("BAR ",            "Bars"),
    ("TAVERN",          "Bars"),
    ("LOUNGE",          "Bars"),
    ("OVERLOOK",        "Bars"),
    ("LIQUOR",          "Bars"),
    ("TURTLE BAY",      "Bars"),
    ("TIKI CHICK",      "Bars"),
    ("DUBLIN HOUSE",    "Bars"),
    # Laundry
    ("LAUNDRY",         "Laundry"),
    ("LAUNDROMAT",      "Laundry"),
    # Furniture
    ("IKEA",            "Furniture"),
    ("WAYFAIR",         "Furniture"),
    ("WEST ELM",        "Furniture"),
    # Haircut
    ("BARBER",          "Haircut"),
    ("SALON",           "Haircut"),
    ("SUPERCUTS",       "Haircut"),
    ("GREAT CLIPS",     "Haircut"),
    # Gas
    ("SHELL",           "Gas"),
    ("BP ",             "Gas"),
    ("EXXON",           "Gas"),
    ("CHEVRON",         "Gas"),
    ("MOBIL",           "Gas"),
    ("SUNOCO",          "Gas"),
    ("GAS STATION",     "Gas"),
    # Groceries / food specialty
    ("PEPPER PALACE",   "Groceries"),
    ("FRESH FOOD",      "Groceries"),
    ("WHOLE EARTH",     "Groceries"),
    # Celsius drink
    ("CELSIUS",         "Celcius"),
    # Rent
    ("CLICKPAY",        "Rent"),
    ("ZEGO",            "Rent"),
    # Tax
    ("TAX",             "Tax Guy"),
    # Transfers / P2P payments
    ("AMAZON RETAIL",   "Transfers"),
    ("PAYPAL",          "Transfers"),
    ("APPLE CASH",      "Transfers"),
    ("APPLE PAY",       "Transfers"),
    ("VENMO",           "Transfers"),
    ("CASH APP",        "Transfers"),
    ("CASHAPP",         "Transfers"),
]

def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a") as f:
        f.write(line + "\n")

def load_config():
    if not CONFIG.exists():
        log("ERROR: config.json not found. Run setup first.")
        sys.exit(1)
    with open(CONFIG) as f:
        return json.load(f)

def load_processed():
    if PROCESSED.exists():
        with open(PROCESSED) as f:
            return set(json.load(f))
    return set()

def save_processed(ids):
    with open(PROCESSED, "w") as f:
        json.dump(list(ids), f)

def categorize(merchant: str) -> str:
    merchant_upper = merchant.upper()
    for keyword, category in CATEGORY_MAP:
        if keyword.upper() in merchant_upper:
            return category
    return "Misc."

def categorize_income(merchant: str) -> str:
    m = merchant.upper()
    if any(k in m for k in ["PAYROLL", "HAVENLY", "SALARY", "DIRECT DEP"]):
        return "Work"
    if any(k in m for k in ["IRS", "TAX REF", "CASTTAXRFD", "NYSTTAXRFD", "TAX REFUND"]):
        return "Tax Return"
    if any(k in m for k in ["ZELLE", "APPLE CASH", "VENMO", "CASHAPP", "CASH APP"]):
        return "Friends"
    if any(k in m for k in ["CLIENT", "INVOICE", "FREELANCE"]):
        return "Clients"
    return "Work"

def parse_amex_email(body: str, subject: str, email_date: str) -> list[dict]:
    """Parse Amex transaction alert emails."""
    transactions = []

    # Pattern 1: fraud alert / standard alert
    # e.g. "$103.20 purchase on 09/27/2025 at PAYPAL *GABRIELLALYNAM"
    p1 = re.findall(
        r'\$([\d,]+\.\d{2})\s+(?:purchase|charge|transaction)\s+on\s+(\d{2}/\d{2}/\d{4})\s+at\s+([^\n\r<"]+?)(?:\s+on your|\s+on\s+your|\s*Card|\s*\.|$)',
        body, re.IGNORECASE
    )
    for amt_str, date_str, merchant in p1:
        try:
            amt = float(amt_str.replace(",", ""))
            txn_date = datetime.strptime(date_str, "%m/%d/%Y").date()
            transactions.append({"date": txn_date, "amount": amt, "merchant": merchant.strip(), "type": "Expense"})
        except ValueError:
            pass

    # Pattern 2: "A charge of $XX.XX has been made at MERCHANT"
    p2 = re.findall(
        r'[Aa]\s+(?:charge|purchase)\s+of\s+\$([\d,]+\.\d{2})\s+(?:has been made|was made)\s+at\s+([^\n\r<"]+?)(?:\s+on\s+(\d{2}/\d{2}/\d{4})|\s+on your|\s*\.)',
        body, re.IGNORECASE
    )
    for amt_str, merchant, date_str in p2:
        try:
            amt = float(amt_str.replace(",", ""))
            if date_str:
                txn_date = datetime.strptime(date_str, "%m/%d/%Y").date()
            else:
                txn_date = datetime.strptime(email_date, "%a, %d %b %Y %H:%M:%S %z").date()
            transactions.append({"date": txn_date, "amount": amt, "merchant": merchant.strip(), "type": "Expense"})
        except (ValueError, Exception):
            pass

    # Pattern 3: "Large Purchase Approved" format
    # e.g. "AMAZON MARKETPLACE PAYMENTS $66.69* Mon, Apr 6, 2026"
    p3 = re.findall(
        r'([A-Z][A-Z0-9 \*&/\'\-]{3,}?)\s+\$([\d,]+\.\d{2})\*?\s+\w+,\s+([A-Z][a-z]+\s+\d+,\s+\d{4})',
        body
    )
    for merchant, amt_str, date_str in p3:
        try:
            amt = float(amt_str.replace(",", ""))
            txn_date = datetime.strptime(date_str.strip(), "%b %d, %Y").date()
            transactions.append({"date": txn_date, "amount": amt, "merchant": merchant.strip(), "type": "Expense"})
        except ValueError:
            pass

    # Pattern 4: subject line "Transaction: $XX.XX at MERCHANT"
    if not transactions:
        p4 = re.search(r'Transaction[:\s]+\$([\d,]+\.\d{2})\s+at\s+(.+)', subject, re.IGNORECASE)
        if p4:
            try:
                amt = float(p4.group(1).replace(",", ""))
                merchant = p4.group(2).strip()
                txn_date = date.today()
                transactions.append({"date": txn_date, "amount": amt, "merchant": merchant, "type": "Expense"})
            except ValueError:
                pass

    return transactions

def parse_wellsfargo_email(body: str, subject: str, email_date: str) -> list[dict]:
    """Parse Wells Fargo transaction alert emails."""
    transactions = []

    # Pattern 1: "A $XX.XX purchase was made with your Wells Fargo Card at MERCHANT on MM/DD/YYYY"
    p1 = re.findall(
        r'[Aa]\s+\$([\d,]+\.\d{2})\s+(?:purchase|charge|debit|transaction)\s+(?:was made|occurred)\s+(?:with|on|at|for|using)\s+your\s+Wells\s+Fargo\s+(?:Debit\s+|Credit\s+)?Card(?:\s+ending\s+in\s+\d{4})?\s+(?:at|for)\s+([^\n\r<"]+?)(?:\s+on\s+(\d{2}/\d{2}/\d{4})|\s*\.|\s*$)',
        body, re.IGNORECASE
    )
    for amt_str, merchant, date_str in p1:
        try:
            amt = float(amt_str.replace(",", ""))
            if date_str:
                txn_date = datetime.strptime(date_str, "%m/%d/%Y").date()
            else:
                txn_date = date.today()
            transactions.append({"date": txn_date, "amount": amt, "merchant": merchant.strip(), "type": "Expense"})
        except ValueError:
            pass

    # Pattern 2: "Your card ending in XXXX was used for $XX.XX at MERCHANT"
    p2 = re.findall(
        r'[Yy]our\s+(?:Wells\s+Fargo\s+)?(?:card|debit|credit)(?:\s+card)?(?:\s+ending\s+in\s+\d{4})?\s+was\s+used\s+for\s+(?:a\s+(?:purchase\s+of\s+)?)?\$([\d,]+\.\d{2})\s+at\s+([^\n\r<"]+?)(?:\s+on\s+(\d{2}/\d{2}/\d{4})|\s*\.|\s*$)',
        body, re.IGNORECASE
    )
    for amt_str, merchant, date_str in p2:
        try:
            amt = float(amt_str.replace(",", ""))
            if date_str:
                txn_date = datetime.strptime(date_str, "%m/%d/%Y").date()
            else:
                txn_date = date.today()
            transactions.append({"date": txn_date, "amount": amt, "merchant": merchant.strip(), "type": "Expense"})
        except ValueError:
            pass

    # Pattern 3: "Purchase amount XX.XX USD / Merchant details at MERCHANT / Date MM/DD/YYYY"
    # e.g. WF online/phone/mail purchase alerts
    p3_amt = re.search(r'Purchase amount\s+([\d,]+\.\d{2})\s+USD', body, re.IGNORECASE)
    p3_merchant = re.search(r'Merchant details\s+at\s+([^\n\r<"]+?)(?:\s+in\s+[\+\d]|\s*$)', body, re.IGNORECASE)
    p3_date = re.search(r'Date\s+(\d{2}/\d{2}/\d{4})', body, re.IGNORECASE)
    if p3_amt and p3_merchant:
        try:
            amt = float(p3_amt.group(1).replace(",", ""))
            merchant = p3_merchant.group(1).strip()
            if p3_date:
                txn_date = datetime.strptime(p3_date.group(1), "%m/%d/%Y").date()
            else:
                txn_date = date.today()
            transactions.append({"date": txn_date, "amount": amt, "merchant": merchant, "type": "Expense"})
        except ValueError:
            pass

    # Pattern 4: deposit / Zelle received
    dep = re.search(r'\$([\d,]+\.\d{2})\s+is now in account', body, re.IGNORECASE)
    if dep:
        try:
            amt = float(dep.group(1).replace(",", ""))
            depositor = ""
            dep2 = re.search(r'Deposited by:\s*([^\n\r]+)', body)
            if dep2:
                depositor = dep2.group(1).strip()
            # Zelle received
            zelle = re.search(r'(\S+)\s+sent you\s+\$([\d,]+\.\d{2})', body, re.IGNORECASE)
            if zelle:
                depositor = zelle.group(1)
                amt = float(zelle.group(2).replace(",", ""))
            transactions.append({"date": date.today(), "amount": amt, "merchant": depositor or "Deposit", "type": "Income"})
        except ValueError:
            pass

    return transactions

def get_email_body(msg) -> str:
    """Extract plain text or HTML body from email."""
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain":
                try:
                    body += part.get_payload(decode=True).decode("utf-8", errors="ignore")
                except Exception:
                    pass
            elif ctype == "text/html" and not body:
                try:
                    raw = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                    body += re.sub(r'<[^>]+>', ' ', unescape(raw))
                except Exception:
                    pass
    else:
        try:
            raw = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
            if msg.get_content_type() == "text/html":
                raw = re.sub(r'<[^>]+>', ' ', unescape(raw))
            body = raw
        except Exception:
            pass
    return body

def append_to_xlsx(transactions: list[dict]):
    """Append new transactions to Financial Tracker.xlsx."""
    if not XLSX_PATH.exists():
        log(f"ERROR: {XLSX_PATH} not found")
        return 0

    wb = load_workbook(XLSX_PATH)
    ws = wb["Transactions"]

    # Find next empty row (look for first row with no date in col B)
    next_row = 3
    for row in ws.iter_rows(min_row=3, max_col=2):
        if row[1].value is not None:
            next_row = row[0].row + 1
        else:
            break

    added = 0
    for t in transactions:
        row_num = next_row + added
        # Row #
        ws.cell(row_num, 1, row_num - 2)
        # Date
        ws.cell(row_num, 2, t["date"])
        ws.cell(row_num, 2).number_format = "MM/DD/YYYY"
        # Type
        ws.cell(row_num, 3, t["type"])
        # Category (pre-computed with duplicate check)
        ws.cell(row_num, 4, t.get("category", categorize(t["merchant"])))
        # Amount
        ws.cell(row_num, 5, t["amount"])
        ws.cell(row_num, 5).number_format = "$#,##0.00"
        # Description (merchant name)
        ws.cell(row_num, 6, t["merchant"])
        added += 1

    wb.save(XLSX_PATH)
    return added

def sync():
    cfg = load_config()
    processed = load_processed()

    gmail_user = cfg["gmail_email"]
    gmail_pass = cfg["gmail_app_password"]

    log(f"Connecting to Gmail as {gmail_user}...")
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(gmail_user, gmail_pass)
    except Exception as e:
        log(f"ERROR: Gmail login failed — {e}")
        log("Make sure IMAP is enabled and your App Password is correct in config.json")
        sys.exit(1)

    mail.select("inbox")

    # Only scan emails received on/after alerts were enabled
    since = cfg.get("alerts_enabled_since", "06-Apr-2026")

    # Targeted searches — only subject lines that indicate actual spend alerts
    searches = [
        (f'FROM "alerts@notify.wellsfargo.com" SINCE {since} SUBJECT "purchase"',     "wellsfargo"),
        (f'FROM "alerts@notify.wellsfargo.com" SINCE {since} SUBJECT "charge"',       "wellsfargo"),
        (f'FROM "alerts@notify.wellsfargo.com" SINCE {since} SUBJECT "transaction"',  "wellsfargo"),
        (f'FROM "alerts@notify.wellsfargo.com" SINCE {since} SUBJECT "paid"',         "wellsfargo_deposit"),
        (f'FROM "alerts@notify.wellsfargo.com" SINCE {since} SUBJECT "deposit"',      "wellsfargo_deposit"),
        (f'FROM "alerts@notify.wellsfargo.com" SINCE {since} SUBJECT "Zelle"',        "wellsfargo_deposit"),
        (f'FROM "AmericanExpress@welcome.americanexpress.com" SINCE {since} SUBJECT "transaction"',    "amex"),
        (f'FROM "AmericanExpress@welcome.americanexpress.com" SINCE {since} SUBJECT "charge"',         "amex"),
        (f'FROM "AmericanExpress@welcome.americanexpress.com" SINCE {since} SUBJECT "purchase"',       "amex"),
        (f'FROM "AmericanExpress@welcome.americanexpress.com" SINCE {since} SUBJECT "Large Purchase"', "amex"),
    ]

    # Load existing transactions for duplicate detection
    wb_check = load_workbook(XLSX_PATH)
    ws_check = wb_check["Transactions"]
    existing = {}  # (date, amount, merchant) -> count
    for row in ws_check.iter_rows(min_row=3, values_only=True):
        if row[1] and row[4]:
            key = (str(row[1])[:10], float(row[4]), str(row[5] or "").upper())
            existing[key] = existing.get(key, 0) + 1

    all_transactions = []
    new_processed = set(processed)

    for search_query, bank in searches:
        _, data = mail.search(None, search_query)
        ids = data[0].split()
        if ids:
            log(f"Found {len(ids)} emails matching: {search_query[-40:]}")

        for eid in ids:
            eid_str = f"{bank}_{eid.decode()}"
            if eid_str in processed:
                continue

            _, msg_data = mail.fetch(eid, "(RFC822)")
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            subject = msg.get("Subject", "")
            email_date = msg.get("Date", "")
            body = get_email_body(msg)

            if bank == "amex":
                txns = parse_amex_email(body, subject, email_date)
            else:
                txns = parse_wellsfargo_email(body, subject, email_date)

            if txns:
                for t in txns:
                    key = (str(t["date"]), t["amount"], t["merchant"].upper())
                    # Count how many times this key appears in the sheet vs already queued this run
                    in_sheet = existing.get(key, 0)
                    in_queue = sum(1 for q in all_transactions
                                   if (str(q["date"]), q["amount"], q["merchant"].upper()) == key)
                    if in_sheet > in_queue:
                        log(f"  (skip duplicate: {t['date']} ${t['amount']:.2f} {t['merchant']})")
                        continue
                    cat = categorize_income(t["merchant"]) if t["type"] == "Income" else categorize(t["merchant"])
                    t["category"] = cat
                    log(f"  → {t['date']} | {t['type']} | {cat} | ${t['amount']:.2f} | {t['merchant']}")
                    all_transactions.append(t)
            else:
                log(f"  (no transaction data found in: {subject[:60]})")

            new_processed.add(eid_str)

    mail.logout()

    if all_transactions:
        added = append_to_xlsx(all_transactions)
        log(f"Added {added} transactions to Financial Tracker.xlsx")

        # Also persist to SQLite
        if _DB_AVAILABLE:
            try:
                _db.init_db()
                _db.populate_budgets()
                for t in all_transactions:
                    t_copy = dict(t)
                    t_copy.setdefault("description", t_copy.get("merchant", ""))
                    t_copy["source"] = "sync"
                    _db.insert_transaction(t_copy)
                log(f"Wrote {len(all_transactions)} transactions to SQLite DB.")
            except Exception as e:
                log(f"Warning: SQLite write failed — {e}")
        else:
            log("Note: db.py not found, skipping SQLite write.")
    else:
        log("No new transactions found.")

    save_processed(new_processed)

if __name__ == "__main__":
    log("=" * 50)
    log("Financial Tracker Sync Starting")
    sync()
    log("Done.")
